###### Set TADPOLE DATA path #########
tadpoleD1D2File = '../../../data/TADPOLE_D1_D2.csv'
DEBMFolder = './pyebm/pyebm'

import os
import tempfile
str_exp=os.path.dirname(os.path.realpath(__file__))
os.chdir(str_exp)
# tmp_dir = tempfile.mkdtemp('debm')
os.makedirs(str_exp + '/IntermediateData/')


text_file=open(str_exp + '/IntermediateData/datafile.path',"w+")
text_file.write(tadpoleD1D2File)
text_file.close()

text_file=open(str_exp + '/IntermediateData/pyebm.path',"w+")
text_file.write(DEBMFolder)
text_file.close()


## Data Preparation
os.system('python3 PrepareData.py')

## Feature Selection
import subprocess
subprocess.call ("/usr/bin/Rscript --vanilla ./SelectFeatures.R", shell=True)

## Estimate Disease State with DEBM
os.system('python3 EstimateDiseaseState.py')

## Predict Features at Future Timepoints and Smoothen the Post Timepoints
os.chdir(str_exp)
subprocess.call ("/usr/bin/Rscript --vanilla ./PredictFeatures.R", shell=True)
os.system('python3 Classify.py')

from openpyxl import load_workbook
import pandas as pd
import numpy as np
P1 = pd.read_csv(str_exp+'/IntermediateData/probabilities.csv')
p1 = P1.values
X=pd.read_excel(str_exp + '/TADPOLE_Submission_EMC1.xlsx',sheet_name='ID 1')

X['CN relative probability']=p1[:,0]
X['MCI relative probability']=p1[:,1]
X['AD relative probability']=p1[:,2]
book = load_workbook(str_exp+'/TADPOLE_Submission_EMC1.xlsx')
writer = pd.ExcelWriter(str_exp+'/TADPOLE_Submission_EMC1.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
X.to_excel(writer, "ID 1",index=False)
writer.save()
    
import subprocess
subprocess.call ("/usr/bin/Rscript --vanilla ./PredictADASVentricles.R", shell=True)

O1 = pd.read_csv(str_exp+'/IntermediateData/TADPOLE_Submission_EMC1.csv')
X['ADAS13'] = O1['ADAS13'].copy()
X['Ventricles_ICV']=O1['Ventricles_ICV'].copy()

idx_nanv = np.isnan(X['Ventricles_ICV'])
P=pd.read_csv(str_exp +'/IntermediateData/PredictionMatrix.csv')

SP = P.loc[idx_nanv,'DiseaseState']
str_in=str_exp + '/IntermediateData/AgeCorrectedLongTADPOLE.csv'
D = pd.read_csv(str_in)
Dv=D['Ventricles']/D['ICV_bl']
S=pd.read_csv(str_exp+'/IntermediateData/PatientStages.csv',header=None)
Y = D['Diagnosis'].copy()
idx_diag=np.logical_not(np.isnan(Y))

S = S[idx_diag]
v_impute = np.zeros(SP.shape[0])
for i in range(SP.shape[0]):
    A=np.abs(S.values - SP.values[i])
    idx_nearest=np.argsort(A[:,0])[:200]
    v_impute[i]=np.nanmean(Dv.values[idx_nearest])
    
X.loc[idx_nanv,'Ventricles_ICV'] = v_impute

os.system('python3 EstimateDiseaseState_bootstrap.py')

xx1=np.zeros((O1.shape[0],10))
xx2=np.zeros((O1.shape[0],10))
for i in range(50):
    O2 = pd.read_csv(str_exp + '/IntermediateData/TADPOLE_Submission_EMC1_'+str(i)+'.csv')
    xx1[:,i] = O2['ADAS13'].values
    xx2[:,i] = O2['Ventricles_ICV'].values
se1=np.nanstd(xx1,axis=1)
se2=np.nanstd(xx2,axis=1)

se2[np.isnan(se2)]=np.nanmean(se2)

X['ADAS13 50% CI lower'] = X['ADAS13'] - 0.674*se1
X['ADAS13 50% CI upper'] = X['ADAS13'] + 0.674*se1
X['Ventricles_ICV 50% CI lower'] = X['Ventricles_ICV'] - 0.674*se2
X['Ventricles_ICV 50% CI upper'] = X['Ventricles_ICV'] + 0.674*se2

book1 = load_workbook(str_exp+'/TADPOLE_Submission_EMC1.xlsx')
writer1 = pd.ExcelWriter(str_exp+'/TADPOLE_Submission_EMC1.xlsx', engine='openpyxl') 
writer1.book = book
writer1.sheets = dict((ws.title, ws) for ws in book.worksheets)
X.to_excel(writer1, "ID 1",index=False)
writer1.save() 
